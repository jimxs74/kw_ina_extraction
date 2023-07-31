from openpyxl import load_workbook, Workbook
import pandas as pd

def write_excel(df, sheet_name, filename):
    """
    Writes the given dataframe to an excel file with the given filename and sheet name.
    If the sheet already exists in the file, the data in the sheet will be overwritten.
    """
    try:
        book = load_workbook(filename)  # Load the existing workbook
    except FileNotFoundError:
        book = Workbook()  # If the file doesn't exist, create a new workbook

    writer = pd.ExcelWriter(filename, engine='openpyxl')
    writer.book = book
    
    if sheet_name in book.sheetnames:  # If sheet already exists, delete it
        idx = book.sheetnames.index(sheet_name)
        sheet = book[sheet_name]
        book.remove(sheet)
        writer.sheets = {ws.title:ws for ws in book.worksheets}
        
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    writer.save()