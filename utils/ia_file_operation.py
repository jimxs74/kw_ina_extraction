from openpyxl import load_workbook, Workbook
import pandas as pd
import tempfile

def write_excel(df, sheet_name, filename):
    """
    Writes the given dataframe to an excel file with the given filename and sheet name.
    If the sheet already exists in the file, the data in the sheet will be overwritten.
    """
    try:
        book = load_workbook(filename)  # Load the existing workbook
    except FileNotFoundError:
        book = Workbook()  # If the file doesn't exist, create a new workbook

    writer = pd.ExcelWriter(filename, engine='openpyxl')
    writer.book = book
    
    if sheet_name in book.sheetnames:  # If sheet already exists, delete it
        #idx = book.sheetnames.index(sheet_name)
        sheet = book[sheet_name]
        book.remove(sheet)
        #writer.sheets = {ws.title:ws for ws in book.worksheets}
        
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    writer.save()

'''
def write_excel(df, sheet_name, filename):
    """
    Writes the given dataframe to an excel file with the given filename and sheet name.
    If the sheet already exists in the file, the data in the sheet will be overwritten.
    """
    try:
        book = load_workbook(filename)  # Load the existing workbook
        if sheet_name in book.sheetnames:  # If sheet already exists, delete it
            sheet = book[sheet_name]
            book.remove(sheet)
            
            # Save the modified book to a temporary file
            with tempfile.NamedTemporaryFile(delete=False) as tmp:
                book.save(tmp.name)
                tmp_filename = tmp.name

            # Load the modified book from the temporary file
            writer = pd.ExcelWriter(tmp_filename, engine='openpyxl')
        else:
            writer = pd.ExcelWriter(filename, engine='openpyxl')
            writer.book = book

    except FileNotFoundError:
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        book = Workbook()
        writer.book = book

    df.to_excel(writer, sheet_name=sheet_name, index=False)
    writer.save()

    # If we used a temporary file, move it to replace the original file
    if 'tmp_filename' in locals():
        import shutil
        shutil.move(tmp_filename, filename)

'''

def save_df_to_excel(df, output_filename, worksheet_name):
    """
    Save a pandas DataFrame to an Excel file.
    """
    with pd.ExcelWriter(output_filename) as writer:
        df.to_excel(writer, sheet_name=worksheet_name, index=False)